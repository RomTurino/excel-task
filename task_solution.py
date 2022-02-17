
import os
import math
import re
from datetime import datetime, timedelta
from openpyxl.reader.excel import load_workbook

file = 'task_support.xlsx'
if file not in os.listdir():
    raise Exception('Я не могу продолжить работу, поместите файл task_support.xlsx в одну директорию с текущим кодом')
wb = load_workbook(filename = file)
sheet = wb['Tasks']




#Сколько четных чисел в этом столбце?

def how_many_even():
    '''
    timeit: 1000 loops, best of 5: 2.41 msec per loop
    cProfile: 11980 function calls in 0.011 seconds
    '''
    even_counter = 0
    for i in range(3,1001):
        number = sheet[f'B{i}'].value
        if number%2==0:
            even_counter+=1
    return even_counter

#Сколько простых чисел в этом столбце?

def eratosthenes(n):
    arr = list(range(n + 1))
    arr[1] = 0
    i = 2
    while i < n + 1:
        if arr[i] != 0:
            j = i + i
            while j <= n:
                arr[j] = 0
                j = j + i
        i += 1
    arr = set(arr)
    arr.remove(0)
    list(arr).sort()
    return arr
def is_prime(num):
    for i in range(2,round(math.sqrt(num)) + 1):
        if num%i==0:
            return False
    return True

def how_many_primes():
    '''
    timeit: 1000 loops, best of 5: 1.44 msec per loop
    cProfile: 3013 function calls in 0.015 seconds
    '''
    max_cell = int(sheet['C1002'].value)
    prime_nums=eratosthenes(max_cell)
    prime_counter = 0
    for i in range(3,1001):
        num = sheet.cell(row = i, column=3).value
        if num in prime_nums:
            prime_counter+=1
    return prime_counter
def how_many_primes2():
    '''
    timeit: 1000 loops, best of 5: 3.31 msec per loop
    cProfile: 14974 function calls in 0.011 seconds
    '''
    prime_counter=0
    for i in range(3,1001):
        number = sheet[f'C{i}'].value
        if is_prime(number):
            prime_counter+=1
    return prime_counter


#Сколько чисел, меньших 0.5 в этом столбце?

def how_many_less_than_half():
    '''
    timeit: 1000 loops, best of 5: 5.6 msec per loop
    cProfile: 21192 function calls (21190 primitive calls) in 0.018 seconds
    '''
    less_counter = 0
    less_list='01234'
    for i in range(3,1001):
        num = re.sub(r'[0\s.,]','',sheet[f'D{i}'].value)
        num = re.findall(r'^\d', num).pop()
        if num in less_list:
            less_counter+=1
    return less_counter


#Столько вторников в этом столбце?
def  how_many_tuesdays():
    '''
    timeit:1000 loops, best of 5: 1.03 msec per loop
    cProfile: 2998 function calls in 0.002 seconds
    '''
    tuesday_counter=0
    for i in range(3,1001):
        date_field = sheet.cell(row = i, column=5).value
        if 'Tue' in date_field:
            tuesday_counter+=1
    return tuesday_counter


#Сколько вторников в этом столбце?
def many_tuesday():
    '''
    timeit: 1000 loops, best of 5: 13.1 msec per loop
    cProfile: 37928 function calls in 0.045 seconds
    '''
    tuesday_counter = 0
    for i in range(3, 1001):
        date_time_str = sheet[f'F{i}'].value
        date_time_obj = datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S.%f')
        if date_time_obj.isoweekday() == 2:
            tuesday_counter+=1
    return tuesday_counter

#Сколько последних вторников месяца в этом столбце?

def how_many_last_tuesday():
    '''
    timeit: 1000 loops, best of 5: 19.2 msec per loop
    cProfile: 28338 function calls in 0.043 seconds
    '''
    tuesday_counter = 0
    #tuesday_list = []
    for i in range(3,1001):
        date_field = sheet.cell(row = i, column=7).value
        date_time_obj = datetime.strptime(date_field, '%m-%d-%Y')
        #worst_case: February(2), 23d
        if date_time_obj.day >=23:
            check_tuesday = date_time_obj.day
            month = date_time_obj.month
            found_tuesday = None
            while month == date_time_obj.month:
                if date_time_obj.isoweekday() == 2:
                    found_tuesday=date_time_obj.day
                date_time_obj+=timedelta(days=1)
            if found_tuesday == check_tuesday:
                tuesday_counter+=1
                #tuesday_list.append(date_field)
    return tuesday_counter #,tuesday_list

def main():
    def question():
        for cell in sheet['B2':'G2']:
            for title in cell:
                yield title.value
    task = question()
    print(task.__next__(), how_many_even())
    print(task.__next__(), how_many_primes())
    print(task.__next__(), how_many_less_than_half())
    print(task.__next__(), how_many_tuesdays())
    print(task.__next__(), many_tuesday())
    print(task.__next__(), how_many_last_tuesday())
main()
